import openpyxl as xl

def find_cell(sheet , max_row , max_col , heading):
    for i in range(1 , max_row + 1):
        for j in range(1 , max_col + 1):
            if sheet.cell(row = i , column = j).value == heading:
                return i , j
            
def process_files(files , mru_class):
    for file in files:
        #result = {"MPR" : [0 , 0]  , "QPR" : [0 , 0]}
        result = {"A" : [0 , 0] , "I" : [0 , 0] , "OTH_MONTHLY" : [0 , 0]  , "QPR" : [0 , 0]}
        sheet = xl.load_workbook(file).active
        max_row , max_col = sheet.max_row , sheet.max_column
        mru_row , mru_col = find_cell(sheet , max_row , max_col , "Billing Portion")
        con_row , con_col = find_cell(sheet , max_row , max_col , "Live & Temp Disc Con")
        inv_row , inv_col = find_cell(sheet , max_row , max_col , "Invoice Generated")

        for i in range(1 , max_row + 1):
            mru = sheet.cell(row = i , column = mru_col).value
            if mru is not None:
                con = sheet.cell(row = i , column = con_col).value
                inv = sheet.cell(row = i , column = inv_col).value
                if con is None : con = 0   #where con count value is blank , consider it as ZERO
                if inv is None : inv = 0   #where invoice count value is blank , consider it as ZERO

                if mru.strip()[-3 : ] == 'MPR':
                    if mru.strip() in mru_class['I'] or mru.strip() in mru_class['S']:
                        result['A'][0] += con
                        result['A'][1] += inv
                    elif mru.strip() in mru_class['A']:
                        result['I'][0] += con
                        result['I'][1] += inv
                    else:
                        result['OTH_MONTHLY'][0] += con
                        result['OTH_MONTHLY'][1] += inv
                elif mru.strip()[-3 : ] == 'QPR':
                    result[mru.strip()[-3:]][0] += con
                    result[mru.strip()[-3:]][1] += inv
                    
                
        print(result)

def mru_wise_class_creation(files):
    mru_class = {}
    sheet = xl.load_workbook(files).active
    max_row , max_col = sheet.max_row , sheet.max_column

    for i in range(1 , max_row + 1):
        mru = sheet.cell(row = i , column = 1).value
        base_class = sheet.cell(row = i , column = 2).value
        mru_class.setdefault(base_class , [])
        mru_class[base_class].append(mru)

    return mru_class

def main():
    print("Hello . Rename CCC wise excels as 103.xlsx , 201.xlsx etc.")
    bill_files = ['103.xlsx' , '201.xlsx' , '202.xlsx' , '208.xlsx' , '300.xlsx' , '301.xlsx' , '401.xlsx' , '402.xlsx']
    mru_files = 'mru_wise_class.xlsx'

    mru_class = mru_wise_class_creation(mru_files)
    process_files(bill_files , mru_class)
    
if __name__ == '__main__':
    main()
